import os
from openpyxl import load_workbook
import csv

# IM群推组id为32
org_id = 100523
sql = 'insert into ims.org_employee_relation(`id`, `org_id`, `employee_id`, `employee_name`, `type`) values (' \
      '0, 100523, "%employee_id%", "%employee_name%", 2);'
delete_sql = 'delete from ims.org_employee_relation where org_id = 100523 and employee_id = "%employee_id%";'

user_file_path = "C:\\Users\\user\\Downloads\\yjbadmin_user_info-1208.xlsx"
# names_file_path = "C:\\Users\\user\\Downloads\\零售线名册-截止930.xlsx"
names_file_path = "C:\\Users\\user\\Downloads\\人员名册1130.xlsx"


def start():
    l = {};
    m = {}
    wb = load_workbook(user_file_path)
    sheet = wb.worksheets[0]
    for i in range(1, sheet.max_row + 1):
        name = sheet.cell(row=i, column=4).value
        ehr_account = sheet.cell(row=i, column=19).value
        # try:
        #     ehr_account = int(sheet.cell(row=i, column=19).value)
        # except Exception as e:
        #     try:
        #         ehr_account = int(sheet.cell(row=i, column=18).value)
        #     except Exception as e2:
        #         print("line " + str(i) + "with no ehr ")
        if len(str(ehr_account)) > 10:
              ehr_account = sheet.cell(row=i, column=18).value
        # if ehr_account is None:
        #     continue
        id = sheet.cell(row=i, column=1).value
        # print(name)
        # print(ehr_account)
        l[str(ehr_account)] = name
        m[str(ehr_account)] = id
    # print(l)

    wb2 = load_workbook(names_file_path)
    sheet2 = wb2.worksheets[0]
    for i in range(2, sheet2.max_row + 1):
        ehr_account = sheet2.cell(row=i, column=3).value
        if ehr_account is None:
            # print("line " + str(i) + " is empty")
            continue
        name = l.get(str(ehr_account))
        if name is None:
            # print("ehr " + str(ehr_account) + " has no name")
            continue
        id = m.get(str(ehr_account))
        if id is None:
            # print("ehr " + str(ehr_account) + " has no id")
            continue
        ds = delete_sql.replace("%employee_id%", str(id))
        print(ds)
        this_sql = sql.replace("%employee_id%", str(id))
        this_sql = this_sql.replace("%employee_name%", str(name))
        print(this_sql)


start()
