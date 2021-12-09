import os
from openpyxl import load_workbook

file_path = "C:\\Users\\user\\Downloads\\定投计划书\\fund2.xlsx"

wb = load_workbook(file_path)
sheet = wb.worksheets[0]
line_start = 3
month = "202111"
replace_flag_enable = "1"
sql = "INSERT INTO `ims`.`ismp_plan_auto_fund` (`id`, `fund_code`, `fund_name`, `month`, `replace_fund_code`, `replace_fund_flag`, `purchase_fund_manul`, `ta_code`) VALUES ('0', 'fundCode', 'fundName', 'monthReplace', 'replaceFundCode', 'replaceFundFlag', '', 'taCode');"

for i in range(line_start, sheet.max_row + 1):
    ta = sheet.cell(row=i, column=1).value
    fund_code = sheet.cell(row=i, column=2).value
    # print(fund_code)
    # if (fund_code == "优选定投基金清单（申购费一折）"):
    #     break
    fund_name = sheet.cell(row=i, column=3).value
    replace_fund_code = sheet.cell(row=i, column=15).value
    replace_fund_flag = replace_flag_enable
    if replace_fund_code is None:
        replace_fund_flag = ""
        replace_fund_code = ""
    else:
        replace_fund_code = str(replace_fund_code).zfill(6)
    if fund_name is None:
        continue
    this_sql = sql.replace("fundCode", str(fund_code).zfill(6))
    this_sql = this_sql.replace("fundName", str(fund_name))
    this_sql = this_sql.replace("monthReplace", str(month))
    this_sql = this_sql.replace("taCode", str(ta))
    this_sql = this_sql.replace("replaceFundCode", replace_fund_code)
    this_sql = this_sql.replace("replaceFundFlag", str(replace_fund_flag))

    print(this_sql)